"""
Historical data importer.

Two sources, each writing into its InfluxDB bucket so the microclimate model has
years of paired forecast-vs-measured data immediately:

  python import_historical.py --ecowitt export.xlsx   (or export.csv)
      Parse an Ecowitt export (web/app "Download"), either CSV or XLSX, into
      bucket `weather`. Column names and units vary between export versions, so
      the parser maps columns by fuzzy name and reads the unit from the header
      (℃/℉, km/h/mph, hPa/inHg, mm/in), converting everything to metric.
      XLSX requires openpyxl (already in requirements.txt).

  python import_historical.py --openmeteo --start 2023-06-01 --end 2026-06-01
      Pull the archived forecast for the location from Open-Meteo's
      Historical Forecast API into bucket `forecasts` (tagged lead_days=1 so it
      pairs with the measured weather in forecast.microclimate).

Run inside the backend container or any environment with the .env values set.
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
from datetime import date, datetime, timezone
from typing import Any

import httpx

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None

from influxdb_client import Point

import config
import db
import settings_store

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger("import")

HIST_FORECAST_URL = "https://historical-forecast-api.open-meteo.com/v1/forecast"


# ── unit conversions ───────────────────────────────────────────────────────
def f_to_c(v): return (v - 32.0) * 5.0 / 9.0
def mph_to_kmh(v): return v * 1.609344
def inhg_to_hpa(v): return v * 33.8638866667
def inch_to_mm(v): return v * 25.4


# Map fuzzy header text -> (field name, kind for unit handling).
# kind: "temp" | "speed" | "pressure" | "rain" | "plain"
COLUMN_MAP = [
    (r"outdoor\s*temp", "temperature_outdoor", "temp"),
    (r"indoor\s*temp", "temperature_indoor", "temp"),
    (r"feels?\s*like", "temperature_feels_like", "temp"),
    (r"dew\s*point", "dewpoint", "temp"),
    (r"outdoor\s*humid", "humidity_outdoor", "plain"),
    (r"indoor\s*humid", "humidity_indoor", "plain"),
    (r"(gust|wind\s*gust)", "wind_gust", "speed"),
    (r"wind\s*direction", "wind_direction", "plain"),
    (r"wind(\s*speed)?\b", "wind_speed", "speed"),
    (r"(rel(ative)?\s*press)", "pressure_relative", "pressure"),
    (r"(abs(olute)?\s*press)", "pressure_absolute", "pressure"),
    (r"rain\s*rate", "rain_rate", "rain"),
    (r"(daily\s*rain|rain.*daily|^rain\b)", "rain_daily", "rain"),
    (r"solar\s*rad", "solar_radiation", "plain"),
    (r"\buvi?\b", "uv_index", "plain"),
]

TIME_FORMATS = [
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%d.%m.%Y %H:%M", "%Y-%m-%dT%H:%M",
]


def _localize(naive: datetime) -> datetime:
    """Treat a naive datetime as station-local time and return it in UTC."""
    if naive.tzinfo is not None:
        return naive.astimezone(timezone.utc)
    if ZoneInfo is not None:
        try:
            return naive.replace(tzinfo=ZoneInfo(config.TIMEZONE)).astimezone(timezone.utc)
        except Exception:  # noqa: BLE001
            pass
    return naive.replace(tzinfo=timezone.utc)


def _parse_time(raw: Any) -> datetime | None:
    # XLSX gives real datetime objects; CSV gives strings.
    if isinstance(raw, datetime):
        return _localize(raw)
    if raw is None:
        return None
    raw = str(raw).strip().strip('"')
    if not raw:
        return None
    if raw.isdigit():  # epoch seconds
        return datetime.fromtimestamp(int(raw), tz=timezone.utc)
    naive = None
    for fmt in TIME_FORMATS:
        try:
            naive = datetime.strptime(raw, fmt)
            break
        except ValueError:
            naive = None
    if naive is None:
        try:
            naive = datetime.fromisoformat(raw)
        except ValueError:
            return None
    return _localize(naive)


def _unit_of(header: str) -> str:
    """Lowercased unit token from the header's parentheses, e.g. '℃', 'mph'."""
    m = re.search(r"\(([^)]*)\)", str(header))
    return (m.group(1) if m else "").lower().replace(" ", "")


def _cell_value(cell: Any) -> float | None:
    """Coerce a CSV string or native XLSX cell to a float, or None if empty/invalid."""
    if cell is None:
        return None
    if isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).strip()
    if s in ("", "--", "---", "null", "None", "N/A", "--.-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _convert(kind: str, unit: str, value: float) -> float:
    if kind == "temp":
        return round(f_to_c(value), 2) if ("℉" in unit or unit in ("f", "°f", "degf")) else round(value, 2)
    if kind == "speed":
        return round(mph_to_kmh(value), 2) if "mph" in unit else round(value, 2)
    if kind == "pressure":
        return round(inhg_to_hpa(value), 2) if "inhg" in unit or unit == "in" else round(value, 2)
    if kind == "rain":
        return round(inch_to_mm(value), 3) if unit in ("in", "inch", "inches", "in/hr") else round(value, 3)
    return round(value, 3)


def _build_header_index(headers: list[str]) -> dict[int, tuple[str, str, str]]:
    """col index -> (field, kind, unit). First matching pattern wins per column."""
    index: dict[int, tuple[str, str, str]] = {}
    used_fields: set[str] = set()
    for i, h in enumerate(headers):
        hl = str(h).lower()
        for pattern, field, kind in COLUMN_MAP:
            if field in used_fields:
                continue
            if re.search(pattern, hl):
                index[i] = (field, kind, _unit_of(h))
                used_fields.add(field)
                break
    return index


def _parse_table(headers: list, row_iter) -> list[dict]:
    """
    Shared parser for both CSV and XLSX: given a header list and an iterable of
    row lists, return [{'ts': datetime, 'fields': {...}}]. Cells may be strings
    (CSV) or native values (XLSX).
    """
    col_index = _build_header_index(headers)
    # Time column = first column whose header looks like a date/time.
    time_col = next(
        (i for i, h in enumerate(headers) if re.search(r"time|date", str(h), re.I)), 0
    )
    if not col_index:
        log.warning("no known weather columns recognised in header: %s", headers)

    rows = []
    for raw in row_iter:
        if not raw or len(raw) <= time_col:
            continue
        ts = _parse_time(raw[time_col])
        if ts is None:
            continue
        fields = {}
        for i, (field, kind, unit) in col_index.items():
            if i >= len(raw):
                continue
            val = _cell_value(raw[i])
            if val is None:
                continue
            fields[field] = _convert(kind, unit, val)
        if fields:
            rows.append({"ts": ts, "fields": fields})
    return rows


def rows_from_ecowitt(path: str) -> list[dict]:
    """
    Parse an Ecowitt export into [{'ts', 'fields'}]. Dispatches by extension:
    .xlsx/.xlsm via openpyxl, everything else as CSV.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError(
                "openpyxl wird zum Lesen von XLSX benötigt: pip install openpyxl"
            ) from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                headers = list(next(it))
            except StopIteration:
                return []
            return _parse_table(headers, (list(r) for r in it))
        finally:
            wb.close()

    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        try:
            headers = next(reader)
        except StopIteration:
            return []
        return _parse_table(headers, reader)


# Backward-compatible alias (CSV + XLSX both supported now).
rows_from_ecowitt_csv = rows_from_ecowitt


def import_ecowitt(path: str, site_id: str) -> int:
    rows = rows_from_ecowitt(path)
    points = [
        Point("station").time(r["ts"]).tag("source", "import").tag("site_id", site_id)
        for r in rows
    ]
    for p, r in zip(points, rows):
        for k, v in r["fields"].items():
            p.field(k, v)
    _write_batched(config.BUCKET_WEATHER, points)
    log.info("imported %d weather rows from %s (site=%s)", len(points), path, site_id)
    return len(points)


# ── Open-Meteo historical forecast ─────────────────────────────────────────
DAILY_VARS = [
    "weathercode", "temperature_2m_max", "temperature_2m_min", "precipitation_sum",
    "precipitation_probability_max", "windspeed_10m_max", "windgusts_10m_max",
    "winddirection_10m_dominant", "sunshine_duration",
]


def fetch_openmeteo_history(start: str, end: str) -> dict:
    params = {
        "latitude": config.LATITUDE,
        "longitude": config.LONGITUDE,
        "start_date": start,
        "end_date": end,
        "timezone": config.TIMEZONE,
        "wind_speed_unit": "kmh",
        "daily": ",".join(DAILY_VARS),
    }
    resp = httpx.get(HIST_FORECAST_URL, params=params, timeout=60)
    resp.raise_for_status()
    return resp.json()


def daily_points_from_openmeteo(raw: dict, site_id: str) -> list[Point]:
    daily = raw.get("daily") or {}
    times = daily.get("time") or []
    points = []
    for i, day_iso in enumerate(times):
        target = date.fromisoformat(day_iso)
        sunshine = _at(daily, "sunshine_duration", i)
        fields = {
            "temp_max": _at(daily, "temperature_2m_max", i),
            "temp_min": _at(daily, "temperature_2m_min", i),
            "precip_sum": _at(daily, "precipitation_sum", i),
            "precip_prob": _at(daily, "precipitation_probability_max", i),
            "wind_max": _at(daily, "windspeed_10m_max", i),
            "gust_max": _at(daily, "windgusts_10m_max", i),
            "wind_dir": _at(daily, "winddirection_10m_dominant", i),
            "weathercode": _at(daily, "weathercode", i),
            "sunshine_hours": round(sunshine / 3600.0, 2) if sunshine else None,
        }
        p = (
            Point("forecast_daily")
            .time(datetime(target.year, target.month, target.day, tzinfo=timezone.utc))
            .tag("target_date", day_iso)
            .tag("lead_days", "1")
            .tag("source", "import")
            .tag("site_id", site_id)
        )
        wrote = False
        for k, v in fields.items():
            if v is not None:
                p.field(k, v)
                wrote = True
        if wrote:
            points.append(p)
    return points


def import_openmeteo(start: str, end: str, site_id: str) -> int:
    raw = fetch_openmeteo_history(start, end)
    points = daily_points_from_openmeteo(raw, site_id)
    _write_batched(config.BUCKET_FORECASTS, points)
    log.info("imported %d archived forecast days (%s..%s, site=%s)", len(points), start, end, site_id)
    return len(points)


# ── helpers ────────────────────────────────────────────────────────────────
def _at(group: dict, key: str, idx: int):
    seq = group.get(key)
    return seq[idx] if isinstance(seq, list) and idx < len(seq) else None


def _write_batched(bucket: str, points: list[Point], batch: int = 1000) -> None:
    for i in range(0, len(points), batch):
        db.write_points(bucket, points[i:i + batch])


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Import historical weather / forecast data into InfluxDB.")
    ap.add_argument("--ecowitt", metavar="FILE", help="path to an Ecowitt export (CSV or XLSX)")
    ap.add_argument("--openmeteo", action="store_true", help="fetch Open-Meteo historical forecast archive")
    ap.add_argument("--start", help="start date YYYY-MM-DD (with --openmeteo)")
    ap.add_argument("--end", help="end date YYYY-MM-DD (with --openmeteo)")
    ap.add_argument("--years", type=int, help="convenience: last N years up to today (with --openmeteo)")
    ap.add_argument("--site", help="site_id to tag the data with (default: the configured default site)")
    args = ap.parse_args(argv)

    if not args.ecowitt and not args.openmeteo:
        ap.error("specify --ecowitt CSV and/or --openmeteo (--start/--end or --years)")

    # Resolve which site the imported data belongs to. Must match an endpoint's
    # ?site= so the microclimate model can pair it (default site claims it too).
    site_id = args.site or settings_store.default_site_id()
    if args.site and args.site not in settings_store.site_ids():
        log.warning("site '%s' is not in settings; importing with that tag anyway", args.site)
    log.info("importing into site '%s'", site_id)

    if args.ecowitt:
        import_ecowitt(args.ecowitt, site_id)
    if args.openmeteo:
        start, end = args.start, args.end
        if args.years:
            today = date.today()
            end = today.isoformat()
            try:
                start = today.replace(year=today.year - args.years).isoformat()
            except ValueError:  # Feb 29 edge case
                start = (today.replace(month=2, day=28, year=today.year - args.years)).isoformat()
        if not start or not end:
            ap.error("--openmeteo requires --start and --end, or --years N")
        import_openmeteo(start, end, site_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
